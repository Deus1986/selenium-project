import time
import numpy as np
import pandas as pd
import requests
import talib
from datetime import datetime, timedelta
import warnings
import os
from openpyxl.styles import PatternFill, Font, Alignment

warnings.filterwarnings('ignore')


def get_high_volume_symbols(min_volume=10000000):
    """Получает список монет с высоким объемом (более 10 миллионов)"""
    url = "https://contract.mexc.com/api/v1/contract/ticker"
    try:
        response = requests.get(url, timeout=10)
        data = response.json()
        symbols = []

        if 'data' in data:
            for item in data["data"]:
                if item["amount24"] > min_volume:
                    symbols.append({
                        'symbol': item['symbol'],
                        'volume_24h': item['amount24'],
                        'price_change_percent': float(item['riseFallRate']) * 100,
                        'last_price': float(item['lastPrice'])
                    })

        return sorted(symbols, key=lambda x: x['volume_24h'], reverse=True)
    except Exception as e:
        print(f"Ошибка получения списка монет: {e}")
        return []


class StrongShortPositionFinder:
    """
    Класс для поиска сильных позиций для шорта с детальным описанием условий входа
    ОПТИМИЗИРОВАН ДЛЯ КОРОТКИХ СДЕЛОК (5 минут, до 1 часа)
    """
    
    def __init__(self):
        self.min_confidence = 70  # Минимальная уверенность для сигнала (сбалансировано)
        self.min_rr_ratio = 2.0   # Минимальное соотношение риск/прибыль 1:2.0 (для коротких сделок)
        
    def get_candles(self, symbol, interval="Min5", limit=100):
        """Получение свечных данных для коротких сделок (5 минут)"""
        url = f"https://contract.mexc.com/api/v1/contract/kline/{symbol}"
        params = {"interval": interval, "limit": limit}
        
        try:
            response = requests.get(url, params=params, timeout=15)
            if response.status_code == 200:
                data = response.json()
                return data
        except Exception as e:
            print(f"   ⚠️ Ошибка получения данных для {symbol}: {e}")
        return None
    
    def create_dataframe(self, data):
        """Создание DataFrame из данных MEXC"""
        if not data or not data.get('success') or not data.get('data'):
            return None
        
        raw_data = data['data']
        
        try:
            if isinstance(raw_data, dict):
                required_fields = ['time', 'open', 'close', 'high', 'low', 'vol']
                if not all(field in raw_data for field in required_fields):
                    return None
                
                df = pd.DataFrame({
                    'timestamp': raw_data['time'],
                    'open': raw_data['open'],
                    'high': raw_data['high'],
                    'low': raw_data['low'],
                    'close': raw_data['close'],
                    'volume': raw_data['vol']
                })
                
                numeric_columns = ['open', 'high', 'low', 'close', 'volume']
                for col in numeric_columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                
                df = df.dropna()
                
                if len(df) < 30:  # Уменьшено для 5-минутного таймфрейма
                    return None
                
                # ИСПРАВЛЕНИЕ: Правильная конвертация timestamp
                # MEXC может возвращать timestamp в секундах или миллисекундах
                try:
                    # Проверяем размер timestamp
                    sample_ts = df['timestamp'].iloc[0]
                    if sample_ts > 1e12:  # Миллисекунды
                        df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
                    else:  # Секунды
                        df['datetime'] = pd.to_datetime(df['timestamp'], unit='s')
                except:
                    # Fallback: пробуем оба варианта
                    try:
                        df['datetime'] = pd.to_datetime(df['timestamp'], unit='s')
                    except:
                        df['datetime'] = pd.to_datetime(df['timestamp'], unit='ms')
                
                df = df.sort_values('datetime').reset_index(drop=True)
                
                return df
            else:
                return None
                
        except Exception as e:
            print(f"   ⚠️ Ошибка создания DataFrame: {e}")
            return None
    
    def calculate_indicators(self, df):
        """Расчет технических индикаторов для коротких сделок"""
        if len(df) < 30:  # ИЗМЕНЕНО: меньше данных для 5-минутного таймфрейма
            return df
        
        close = df['close'].values
        high = df['high'].values
        low = df['low'].values
        volume = df['volume'].values
        
        try:
            # Трендовые индикаторы
            df['sma_20'] = talib.SMA(close, timeperiod=20)
            df['sma_50'] = talib.SMA(close, timeperiod=50)
            df['ema_12'] = talib.EMA(close, timeperiod=12)
            df['ema_26'] = talib.EMA(close, timeperiod=26)
            df['ema_50'] = talib.EMA(close, timeperiod=50)
            
            # MACD
            df['macd'], df['macd_signal'], df['macd_hist'] = talib.MACD(close)
            
            # RSI на разных периодах
            df['rsi_14'] = talib.RSI(close, timeperiod=14)
            df['rsi_21'] = talib.RSI(close, timeperiod=21)
            
            # Stochastic
            df['stoch_k'], df['stoch_d'] = talib.STOCH(high, low, close)
            
            # Bollinger Bands
            df['bb_upper'], df['bb_middle'], df['bb_lower'] = talib.BBANDS(close, timeperiod=20)
            
            # Волатильность
            df['atr'] = talib.ATR(high, low, close, timeperiod=14)
            df['natr'] = talib.NATR(high, low, close, timeperiod=14)
            
            # Сила тренда
            df['adx'] = talib.ADX(high, low, close, timeperiod=14)
            df['plus_di'] = talib.PLUS_DI(high, low, close, timeperiod=14)
            df['minus_di'] = talib.MINUS_DI(high, low, close, timeperiod=14)
            
            # Объемы
            df['volume_sma'] = talib.SMA(volume, timeperiod=20)
            df['volume_ratio'] = volume / df['volume_sma']
            
            # Уровни поддержки и сопротивления
            df['resistance'] = df['high'].rolling(window=20).max()
            df['support'] = df['low'].rolling(window=20).min()
            
            # Анализ свечей
            df['candle_body'] = abs(df['close'] - df['open'])
            df['candle_size'] = df['high'] - df['low']
            df['body_ratio'] = np.where(df['candle_size'] > 0, df['candle_body'] / df['candle_size'], 0)
            df['is_red'] = df['close'] < df['open']
            
        except Exception as e:
            print(f"   ⚠️ Ошибка расчета индикаторов: {e}")
        
        return df
    
    def analyze_short_conditions(self, df):
        """Анализ условий для шорта (УЛУЧШЕННЫЕ КРИТЕРИИ ДЛЯ КОРОТКИХ СДЕЛОК)"""
        if len(df) < 30:
            return None
        
        current = df.iloc[-1]
        conditions = []
        score = 0
        
        # 1. RSI перекупленность (важно для шорта)
        rsi_14 = current['rsi_14']
        if rsi_14 >= 75:  # Очень сильная перекупленность
            conditions.append(f"🔴 КРИТИЧЕСКАЯ перекупленность RSI: {rsi_14:.1f}")
            score += 35
        elif rsi_14 >= 70:  # Сильная перекупленность
            conditions.append(f"🔴 Сильная перекупленность RSI: {rsi_14:.1f}")
            score += 30
        elif rsi_14 >= 65:  # Умеренная перекупленность
            conditions.append(f"🟠 Перекупленность RSI: {rsi_14:.1f}")
            score += 20
        elif rsi_14 >= 60:  # Слабая перекупленность
            conditions.append(f"🟡 Умеренная перекупленность RSI: {rsi_14:.1f}")
            score += 10
        else:
            # Не перекуплен - не подходит для шорта, но не отклоняем полностью
            score -= 10
        
        # 2. Медвежий тренд (SMA/EMA)
        if current['sma_20'] < current['sma_50']:
            conditions.append("📉 Медвежий тренд (SMA20 < SMA50)")
            score += 15
        
        if current['ema_12'] < current['ema_26']:
            conditions.append("📉 Медвежий тренд (EMA12 < EMA26)")
            score += 15
        
        # 3. MACD медвежий
        if current['macd'] < current['macd_signal']:
            conditions.append("📉 Медвежий MACD")
            score += 15
        
        # 4. Цена у верхней границы Bollinger Bands
        bb_position = (current['close'] - current['bb_lower']) / (current['bb_upper'] - current['bb_lower'])
        if bb_position >= 0.8:
            conditions.append(f"📊 Цена у верхней границы BB ({bb_position*100:.1f}%)")
            score += 15
        elif bb_position >= 0.7:
            conditions.append(f"📊 Цена близко к верхней границе BB ({bb_position*100:.1f}%)")
            score += 10
        
        # 5. Stochastic перекупленность
        stoch_k = current['stoch_k']
        if stoch_k >= 80:
            conditions.append(f"📊 Stochastic перекуплен: {stoch_k:.1f}")
            score += 10
        
        # 6. Высокий объем при падении
        volume_ratio = current['volume_ratio']
        is_red_candle = current['is_red']
        
        if is_red_candle and volume_ratio > 2.0:
            conditions.append(f"📊 ОЧЕНЬ высокий объем при падении (x{volume_ratio:.2f})")
            score += 25
        elif is_red_candle and volume_ratio > 1.5:
            conditions.append(f"📊 Высокий объем при падении (x{volume_ratio:.2f})")
            score += 20
        elif is_red_candle and volume_ratio > 1.0:
            conditions.append(f"📊 Красная свеча с нормальным объемом x{volume_ratio:.2f}")
            score += 10
        elif is_red_candle and volume_ratio > 0.7:
            conditions.append(f"📊 Красная свеча, объем умеренный x{volume_ratio:.2f}")
            score += 5
        elif is_red_candle:
            # Низкий объем на красной свече
            conditions.append(f"⚠️ Красная свеча, но низкий объем x{volume_ratio:.2f}")
            score -= 5
        else:
            # Зеленая свеча - не хорошо для шорта
            score -= 15
        
        # 7. Сила тренда (ADX) - КРИТИЧЕСКИ ВАЖНО
        adx = current['adx']
        if adx >= 30:  # ИЗМЕНЕНО: повышен порог для сильного тренда
            conditions.append(f"💪 ОЧЕНЬ сильный тренд (ADX: {adx:.1f})")
            score += 20
            if current['minus_di'] > current['plus_di'] + 5:  # НОВОЕ: требуем явное преимущество
                conditions.append("📉 СИЛЬНАЯ медвежья сила (MINUS_DI >> PLUS_DI)")
                score += 15
        elif adx >= 25:
            conditions.append(f"💪 Сильный тренд (ADX: {adx:.1f})")
            score += 10
            if current['minus_di'] > current['plus_di']:
                conditions.append("📉 Медвежья сила тренда (MINUS_DI > PLUS_DI)")
                score += 10
        else:
            # Слабый тренд - плохо для шорта
            score -= 5
        
        # 8. Цена у сопротивления
        resistance = current['resistance']
        current_price = current['close']
        distance_to_resistance = ((resistance - current_price) / current_price) * 100
        if distance_to_resistance <= 1.0:
            conditions.append(f"🎯 Цена у уровня сопротивления ({distance_to_resistance:.2f}%)")
            score += 20
        elif distance_to_resistance <= 2.0:
            conditions.append(f"🎯 Цена близко к сопротивлению ({distance_to_resistance:.2f}%)")
            score += 10
        
        # 9. Серия красных свечей
        red_candles_count = df['is_red'].tail(3).sum()
        if red_candles_count >= 2:
            conditions.append(f"📉 Серия из {red_candles_count} красных свечей")
            score += 10
        
        # 10. Медвежья дивергенция (упрощенная проверка)
        if len(df) >= 20:
            recent_highs = df['high'].tail(10).values
            recent_rsi = df['rsi_14'].tail(10).values
            if len(recent_highs) >= 5 and len(recent_rsi) >= 5:
                price_trend = recent_highs[-1] > recent_highs[-5]
                rsi_trend = recent_rsi[-1] < recent_rsi[-5]
                if price_trend and rsi_trend:
                    conditions.append("📉 Медвежья дивергенция (цена растет, RSI падает)")
                    score += 15
        
        return {
            'conditions': conditions,
            'score': min(score, 100),
            'rsi_14': rsi_14,
            'rsi_21': current['rsi_21'],
            'adx': adx,
            'volume_ratio': volume_ratio,
            'bb_position': bb_position,
            'distance_to_resistance': distance_to_resistance
        }
    
    def calculate_entry_levels(self, df, analysis):
        """Расчет уровней для КОРОТКИХ СДЕЛОК (до 1 часа)"""
        if len(df) < 30:
            return None
        
        current = df.iloc[-1]
        current_price = current['close']
        atr = current['atr']
        resistance = current['resistance']
        
        # ИЗМЕНЕНО: Более узкие уровни для коротких сделок
        # Стоп-лосс: ближе к цене для быстрых выходов
        stop_loss_1 = resistance * 1.003  # На 0.3% выше сопротивления (было 0.5%)
        stop_loss_2 = current_price + (atr * 1.0)  # На 1.0 ATR выше (было 1.5)
        stop_loss = max(stop_loss_1, stop_loss_2)
        
        # ИЗМЕНЕНО: Более реалистичные цели для коротких сделок
        # На 5-минутном графике движения меньше
        # Уровень 1: 1.5 ATR вниз (консервативный для быстрых сделок)
        take_profit_1 = current_price - (atr * 1.5)
        
        # Уровень 2: 2.0 ATR вниз (умеренный)
        take_profit_2 = current_price - (atr * 2.0)
        
        # Уровень 3: 2.5 ATR вниз (агрессивный)
        take_profit_3 = current_price - (atr * 2.5)
        
        # Проверяем соотношение риск/прибыль
        risk = stop_loss - current_price
        reward_1 = current_price - take_profit_1
        reward_2 = current_price - take_profit_2
        reward_3 = current_price - take_profit_3
        
        rr_ratio_1 = reward_1 / risk if risk > 0 else 0
        rr_ratio_2 = reward_2 / risk if risk > 0 else 0
        rr_ratio_3 = reward_3 / risk if risk > 0 else 0
        
        # Если соотношение плохое, корректируем
        if rr_ratio_1 < self.min_rr_ratio:
            # Увеличиваем цель
            take_profit_1 = current_price - (risk * self.min_rr_ratio)
            reward_1 = current_price - take_profit_1
            rr_ratio_1 = reward_1 / risk if risk > 0 else 0
        
        # Процентные изменения
        stop_loss_percent = ((stop_loss / current_price) - 1) * 100
        take_profit_1_percent = ((current_price - take_profit_1) / current_price) * 100
        take_profit_2_percent = ((current_price - take_profit_2) / current_price) * 100
        take_profit_3_percent = ((current_price - take_profit_3) / current_price) * 100
        
        return {
            'entry_price': round(current_price, 6),
            'stop_loss': round(stop_loss, 6),
            'stop_loss_percent': round(stop_loss_percent, 2),
            'take_profit_1': round(take_profit_1, 6),
            'take_profit_1_percent': round(take_profit_1_percent, 2),
            'take_profit_2': round(take_profit_2, 6),
            'take_profit_2_percent': round(take_profit_2_percent, 2),
            'take_profit_3': round(take_profit_3, 6),
            'take_profit_3_percent': round(take_profit_3_percent, 2),
            'risk': round(risk, 6),
            'reward_1': round(reward_1, 6),
            'reward_2': round(reward_2, 6),
            'reward_3': round(reward_3, 6),
            'rr_ratio_1': round(rr_ratio_1, 2),
            'rr_ratio_2': round(rr_ratio_2, 2),
            'rr_ratio_3': round(rr_ratio_3, 2),
            'atr': round(atr, 6),
            'resistance_level': round(resistance, 6)
        }
    
    def determine_entry_timing(self, df):
        """Определение оптимального времени входа (УЛУЧШЕННАЯ ВЕРСИЯ)"""
        current = df.iloc[-1]
        current_time = current['datetime']
        
        # Анализируем последние свечи для определения момента входа
        recommendations = []
        urgency_score = 0
        
        # НОВОЕ: Проверяем последние 3 свечи на подтверждение тренда
        if len(df) >= 3:
            last_3_closes = df['close'].tail(3).values
            red_count = sum(1 for i in range(1, 3) if last_3_closes[i] < last_3_closes[i-1])
            
            if red_count >= 2:
                recommendations.append(f"✅ ПОДТВЕРЖДЕНИЕ: {red_count} из 2 последних свечей красные")
                urgency_score += 2
        
        # Проверяем текущую свечу
        is_red = current['is_red']
        volume_ratio = current['volume_ratio']
        
        if is_red and volume_ratio > 1.5:  # ИЗМЕНЕНО: повышен порог объема
            recommendations.append("✅ ВХОДИТЬ СЕЙЧАС: Красная свеча с высоким объемом")
            urgency_score += 3
        elif is_red and volume_ratio > 1.0:
            recommendations.append("⚠️ ОСТОРОЖНО: Красная свеча, но объем средний")
            urgency_score += 1
        else:
            recommendations.append("❌ НЕ ВХОДИТЬ: Свеча зеленая или низкий объем")
            urgency_score -= 2
        
        # Проверяем RSI (УЖЕСТОЧЕНО)
        rsi = current['rsi_14']
        if rsi >= 75:  # ИЗМЕНЕНО: повышен порог
            recommendations.append("✅ ОТЛИЧНО: RSI в СИЛЬНОЙ перекупленности")
            urgency_score += 3
        elif rsi >= 70:
            recommendations.append("✅ ХОРОШО: RSI в зоне перекупленности")
            urgency_score += 2
        elif rsi >= 65:
            recommendations.append("⚠️ УМЕРЕННО: RSI близок к перекупленности")
            urgency_score += 1
        else:
            recommendations.append("❌ ПЛОХО: RSI слишком низкий для шорта")
            urgency_score -= 1
        
        # Проверяем цену относительно BB (УЖЕСТОЧЕНО)
        bb_position = (current['close'] - current['bb_lower']) / (current['bb_upper'] - current['bb_lower'])
        if bb_position >= 0.9:  # ИЗМЕНЕНО: требуем быть очень близко к верхней границе
            recommendations.append("✅ ОТЛИЧНО: Цена у САМОЙ верхней границы BB")
            urgency_score += 2
        elif bb_position >= 0.8:
            recommendations.append("✅ ХОРОШО: Цена у верхней границы BB")
            urgency_score += 1
        
        # НОВОЕ: Проверяем momentum
        if len(df) >= 5:
            price_5_ago = df['close'].iloc[-5]
            current_price = current['close']
            momentum_5 = ((current_price - price_5_ago) / price_5_ago) * 100
            
            if momentum_5 > 2:  # Сильный рост перед разворотом
                recommendations.append(f"✅ MOMENTUM: Сильный рост {momentum_5:.1f}% за 5 свечей")
                urgency_score += 2
        
        # Определяем срочность
        if urgency_score >= 8:
            entry_urgency = 'КРИТИЧЕСКАЯ - ВХОДИТЬ НЕМЕДЛЕННО!'
        elif urgency_score >= 5:
            entry_urgency = 'ВЫСОКАЯ - Хороший момент'
        elif urgency_score >= 3:
            entry_urgency = 'СРЕДНЯЯ - Можно входить'
        else:
            entry_urgency = 'НИЗКАЯ - Ожидать лучшего момента'
        
        return {
            'current_time': current_time.strftime('%Y-%m-%d %H:%M:%S'),
            'recommendations': recommendations,
            'entry_urgency': entry_urgency,
            'urgency_score': urgency_score
        }
    
    def analyze_symbol(self, symbol):
        """Полный анализ символа для шорта"""
        print(f"\n🔍 Анализируем {symbol} для шорта...")
        
        try:
            # Получаем данные
            # ИЗМЕНЕНО: 5-минутный таймфрейм для коротких сделок
            data = self.get_candles(symbol, "Min5", 100)
            if not data:
                print(f"   ❌ Нет данных для {symbol}")
                return None
            
            df = self.create_dataframe(data)
            if df is None or len(df) < 50:
                print(f"   ❌ Недостаточно данных для {symbol}")
                return None
            
            # Рассчитываем индикаторы
            df = self.calculate_indicators(df)
            
            # Анализируем условия
            analysis = self.analyze_short_conditions(df)
            if not analysis:
                print(f"   ❌ Не удалось проанализировать {symbol}")
                return None
            
            # Проверяем минимальную уверенность
            if analysis['score'] < self.min_confidence:
                print(f"   ⚠️ Низкая уверенность: {analysis['score']}/100 (нужно ≥{self.min_confidence})")
                print(f"      RSI: {analysis['rsi_14']:.1f}, Объем: {analysis['volume_ratio']:.2f}x, ADX: {analysis['adx']:.1f}")
                return None
            
            # Рассчитываем уровни
            levels = self.calculate_entry_levels(df, analysis)
            if not levels:
                print(f"   ❌ Не удалось рассчитать уровни для {symbol}")
                return None
            
            # Проверяем минимальное соотношение R/R
            if levels['rr_ratio_1'] < self.min_rr_ratio:
                print(f"   ⚠️ Низкое соотношение R/R: {levels['rr_ratio_1']}:1 (минимум {self.min_rr_ratio}:1)")
                return None
            
            # Определяем время входа
            timing = self.determine_entry_timing(df)
            
            result = {
                'symbol': symbol,
                'confidence': analysis['score'],
                'conditions': analysis['conditions'],
                'levels': levels,
                'timing': timing,
                'indicators': {
                    'rsi_14': round(analysis['rsi_14'], 1),
                    'rsi_21': round(analysis['rsi_21'], 1),
                    'adx': round(analysis['adx'], 1),
                    'volume_ratio': round(analysis['volume_ratio'], 2),
                    'bb_position': round(analysis['bb_position'] * 100, 1)
                },
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            print(f"   ✅ Найден сильный сигнал! Уверенность: {analysis['score']}/100")
            
            return result
            
        except Exception as e:
            print(f"   ❌ Ошибка анализа {symbol}: {e}")
            return None
    
    def print_signal_details(self, signal):
        """Вывод детальной информации о сигнале"""
        print(f"\n{'='*80}")
        print(f"🎯 СИЛЬНАЯ ПОЗИЦИЯ ДЛЯ ШОРТА: {signal['symbol']}")
        print(f"{'='*80}")
        
        print(f"\n📊 ОБЩАЯ ИНФОРМАЦИЯ:")
        print(f"   • Уверенность сигнала: {signal['confidence']}/100")
        print(f"   • Время анализа: {signal['timestamp']}")
        print(f"   • Текущее время рынка: {signal['timing']['current_time']}")
        print(f"   • Срочность входа: {signal['timing']['entry_urgency']}")
        
        print(f"\n📈 ТЕХНИЧЕСКИЕ ИНДИКАТОРЫ:")
        print(f"   • RSI(14): {signal['indicators']['rsi_14']}")
        print(f"   • RSI(21): {signal['indicators']['rsi_21']}")
        print(f"   • ADX: {signal['indicators']['adx']}")
        print(f"   • Объем: {signal['indicators']['volume_ratio']}x от среднего")
        print(f"   • Позиция в BB: {signal['indicators']['bb_position']}%")
        
        print(f"\n🎯 УСЛОВИЯ ВХОДА ({len(signal['conditions'])}):")
        for i, condition in enumerate(signal['conditions'], 1):
            print(f"   {i}. {condition}")
        
        print(f"\n💰 ТОРГОВЫЕ УРОВНИ:")
        levels = signal['levels']
        print(f"   • Цена входа: {levels['entry_price']} USDT")
        print(f"   • Стоп-лосс: {levels['stop_loss']} USDT (+{levels['stop_loss_percent']}%)")
        print(f"   • Уровень сопротивления: {levels['resistance_level']} USDT")
        print(f"   • ATR: {levels['atr']} USDT")
        
        print(f"\n🎯 ЦЕЛИ ВЫХОДА:")
        print(f"   1. Консервативная цель: {levels['take_profit_1']} USDT (-{levels['take_profit_1_percent']}%)")
        print(f"      → Прибыль: {levels['reward_1']} USDT | R/R: {levels['rr_ratio_1']}:1")
        print(f"   2. Умеренная цель: {levels['take_profit_2']} USDT (-{levels['take_profit_2_percent']}%)")
        print(f"      → Прибыль: {levels['reward_2']} USDT | R/R: {levels['rr_ratio_2']}:1")
        print(f"   3. Агрессивная цель: {levels['take_profit_3']} USDT (-{levels['take_profit_3_percent']}%)")
        print(f"      → Прибыль: {levels['reward_3']} USDT | R/R: {levels['rr_ratio_3']}:1")
        
        print(f"\n⏰ РЕКОМЕНДАЦИИ ПО ВРЕМЕНИ ВХОДА:")
        for i, rec in enumerate(signal['timing']['recommendations'], 1):
            print(f"   {i}. {rec}")
        
        print(f"\n💡 ИНСТРУКЦИЯ ПО ВХОДУ (КОРОТКАЯ СДЕЛКА):")
        print(f"   ⏱️ ВРЕМЯ СДЕЛКИ: до 1 часа (12 свечей по 5 минут)")
        print(f"   1. ВХОД: Открыть SHORT позицию по цене {levels['entry_price']} USDT")
        print(f"   2. СТОП-ЛОСС: Установить на уровне {levels['stop_loss']} USDT")
        print(f"   3. ТЕЙК-ПРОФИТ 1: Закрыть 50% позиции на {levels['take_profit_1']} USDT (~15-20 мин)")
        print(f"   4. ТЕЙК-ПРОФИТ 2: Закрыть еще 30% позиции на {levels['take_profit_2']} USDT (~30-40 мин)")
        print(f"   5. ТЕЙК-ПРОФИТ 3: Закрыть оставшиеся 20% на {levels['take_profit_3']} USDT (~45-60 мин)")
        print(f"   6. РИСК: Максимальный риск на сделку не более 1-2% от депозита")
        print(f"   7. ⚠️ ВАЖНО: Если через 1 час цель не достигнута - закрыть принудительно!")
        
        print(f"\n{'='*80}\n")
    
    def export_to_excel(self, signals, filename=None):
        """Экспорт результатов в Excel таблицу"""
        if not signals:
            print("\n⚠️ Нет данных для экспорта в Excel")
            return None
        
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"short_signals_{timestamp}.xlsx"
        
        # Подготовка данных для основной таблицы
        data_rows = []
        
        for i, signal in enumerate(signals, 1):
            levels = signal['levels']
            indicators = signal['indicators']
            
            # Собираем условия в одну строку
            conditions_text = " | ".join(signal['conditions'][:5])  # Первые 5 условий
            
            # Рекомендации по входу
            timing_recommendations = " | ".join(signal['timing']['recommendations'])
            
            row = {
                '№': i,
                'Символ': signal['symbol'],
                'Уверенность (%)': signal['confidence'],
                'Время анализа': signal['timestamp'],
                'Текущее время рынка': signal['timing']['current_time'],
                'Срочность входа': signal['timing']['entry_urgency'],
                
                # Торговые уровни
                'Цена входа (USDT)': levels['entry_price'],
                'Стоп-лосс (USDT)': levels['stop_loss'],
                'Стоп-лосс (%)': levels['stop_loss_percent'],
                
                # Цели
                'Цель 1 (USDT)': levels['take_profit_1'],
                'Цель 1 (%)': levels['take_profit_1_percent'],
                'R/R Цель 1': levels['rr_ratio_1'],
                
                'Цель 2 (USDT)': levels['take_profit_2'],
                'Цель 2 (%)': levels['take_profit_2_percent'],
                'R/R Цель 2': levels['rr_ratio_2'],
                
                'Цель 3 (USDT)': levels['take_profit_3'],
                'Цель 3 (%)': levels['take_profit_3_percent'],
                'R/R Цель 3': levels['rr_ratio_3'],
                
                # Риски и прибыли
                'Риск (USDT)': levels['risk'],
                'Прибыль 1 (USDT)': levels['reward_1'],
                'Прибыль 2 (USDT)': levels['reward_2'],
                'Прибыль 3 (USDT)': levels['reward_3'],
                
                # Технические данные
                'ATR': levels['atr'],
                'Сопротивление': levels['resistance_level'],
                
                # Индикаторы
                'RSI(14)': indicators['rsi_14'],
                'RSI(21)': indicators['rsi_21'],
                'ADX': indicators['adx'],
                'Объем (x)': indicators['volume_ratio'],
                'Позиция в BB (%)': indicators['bb_position'],
                
                # Условия и рекомендации
                'Условия входа': conditions_text,
                'Рекомендации по времени': timing_recommendations
            }
            
            data_rows.append(row)
        
        # Создаем DataFrame
        df = pd.DataFrame(data_rows)
        
        # Экспорт в Excel
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Основная таблица
                df.to_excel(writer, sheet_name='Сигналы SHORT', index=False)
                
                # Детальная информация по каждому сигналу
                details_rows = []
                for signal in signals:
                    for i, condition in enumerate(signal['conditions'], 1):
                        details_rows.append({
                            'Символ': signal['symbol'],
                            '№ условия': i,
                            'Условие': condition
                        })
                
                if details_rows:
                    df_details = pd.DataFrame(details_rows)
                    df_details.to_excel(writer, sheet_name='Детали условий', index=False)
                
                # Инструкции по входу
                instructions_rows = []
                for signal in signals:
                    levels = signal['levels']
                    instructions_rows.append({
                        'Символ': signal['symbol'],
                        'Инструкция 1': f"ВХОД: Открыть SHORT по цене {levels['entry_price']} USDT",
                        'Инструкция 2': f"СТОП-ЛОСС: Установить на {levels['stop_loss']} USDT",
                        'Инструкция 3': f"ТЕЙК 1 (50%): Закрыть на {levels['take_profit_1']} USDT",
                        'Инструкция 4': f"ТЕЙК 2 (30%): Закрыть на {levels['take_profit_2']} USDT",
                        'Инструкция 5': f"ТЕЙК 3 (20%): Закрыть на {levels['take_profit_3']} USDT",
                        'Инструкция 6': "РИСК: Не более 2% от депозита"
                    })
                
                df_instructions = pd.DataFrame(instructions_rows)
                df_instructions.to_excel(writer, sheet_name='Инструкции', index=False)
                
                # Автоматическая настройка ширины столбцов для всех листов
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # Проходим по всем столбцам и настраиваем ширину
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if cell.value:
                                    # Вычисляем максимальную длину содержимого
                                    cell_length = len(str(cell.value))
                                    if cell_length > max_length:
                                        max_length = cell_length
                            except:
                                pass
                        
                        # Устанавливаем ширину столбца (с небольшим запасом)
                        adjusted_width = min(max_length + 2, 100)  # Максимум 100 символов
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Применяем форматирование для листа "Сигналы SHORT"
                if 'Сигналы SHORT' in writer.sheets:
                    ws = writer.sheets['Сигналы SHORT']
                    
                    # Определяем красный цвет и жирный шрифт для умеренной цели (Цель 2)
                    red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    bold_font = Font(bold=True, color='CC0000')
                    center_alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Находим столбцы с "Цель 2"
                    target_2_columns = []
                    header_row = 1
                    
                    for col_idx, cell in enumerate(ws[header_row], start=1):
                        if cell.value and 'Цель 2' in str(cell.value):
                            target_2_columns.append(col_idx)
                    
                    # Применяем форматирование к столбцам с Целью 2
                    for row_idx in range(2, ws.max_row + 1):  # Начиная со 2-й строки (пропускаем заголовок)
                        for col_idx in target_2_columns:
                            cell = ws.cell(row=row_idx, column=col_idx)
                            cell.fill = red_fill
                            cell.font = bold_font
                            cell.alignment = center_alignment
                    
                    # Форматируем заголовки столбцов с Целью 2
                    header_fill = PatternFill(start_color='CC0000', end_color='CC0000', fill_type='solid')
                    header_font = Font(bold=True, color='FFFFFF')
                    
                    for col_idx in target_2_columns:
                        header_cell = ws.cell(row=header_row, column=col_idx)
                        header_cell.fill = header_fill
                        header_cell.font = header_font
                        header_cell.alignment = center_alignment
            
            print(f"\n✅ Результаты успешно экспортированы в файл: {filename}")
            print(f"📁 Полный путь: {os.path.abspath(filename)}")
            print(f"📊 Ширина столбцов автоматически настроена под содержимое")
            print(f"🔴 Умеренная цель выхода (Цель 2) выделена красным цветом")
            return filename
            
        except Exception as e:
            print(f"\n❌ Ошибка экспорта в Excel: {e}")
            print("💡 Убедитесь, что установлен модуль openpyxl: pip install openpyxl")
            return None


def main_strong_short_search():
    """Основная функция поиска сильных шорт-позиций ДЛЯ КОРОТКИХ СДЕЛОК"""
    print("🚀 ПОИСК СИЛЬНЫХ ПОЗИЦИЙ ДЛЯ ШОРТА (КОРОТКИЕ СДЕЛКИ)")
    print("="*80)
    print("⏱️ ОПТИМИЗИРОВАНО ДЛЯ СДЕЛОК ДО 1 ЧАСА")
    print("="*80)
    print("Критерии поиска:")
    print("   • Таймфрейм: 5 минут (Min5)")
    print("   • Длительность сделки: до 1 часа")
    print("   • Минимальная уверенность: 70/100")
    print("   • Минимальное соотношение R/R: 1:2.0")
    print("   • Объем торгов более 10 миллионов USDT")
    print("   • Рекомендуется RSI ≥ 65 для лучших сигналов")
    print("   • Рекомендуется объем ≥ 1.0x для подтверждения")
    print("="*80)
    
    finder = StrongShortPositionFinder()
    
    # Получаем список монет с высоким объемом (более 10 миллионов)
    symbols_data = get_high_volume_symbols(min_volume=10000000)
    if not symbols_data:
        print("❌ Не удалось получить список монет")
        return []
    
    # Анализируем все монеты с объемом более 10 миллионов
    symbols = [item['symbol'] for item in symbols_data]
    
    print(f"\n📊 Анализируем {len(symbols)} монет с высоким объемом...")
    print(f"⏰ Начало анализа: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    strong_signals = []
    
    for symbol in symbols:
        signal = finder.analyze_symbol(symbol)
        if signal:
            strong_signals.append(signal)
        time.sleep(0.5)  # Пауза между запросами
    
    # Сортируем по уверенности
    strong_signals.sort(key=lambda x: x['confidence'], reverse=True)
    
    # Вывод результатов
    print(f"\n{'='*80}")
    print(f"📊 РЕЗУЛЬТАТЫ ПОИСКА СИЛЬНЫХ ШОРТ-ПОЗИЦИЙ")
    print(f"{'='*80}")
    
    if strong_signals:
        print(f"\n🎯 Найдено {len(strong_signals)} сильных сигналов для шорта:\n")
        
        for i, signal in enumerate(strong_signals, 1):
            print(f"\n{'='*80}")
            print(f"СИГНАЛ #{i}")
            finder.print_signal_details(signal)
        
        # Сводная таблица
        print(f"\n{'='*80}")
        print("📋 СВОДНАЯ ТАБЛИЦА СИГНАЛОВ:")
        print(f"{'='*80}")
        print(f"{'№':<4} {'Символ':<15} {'Уверенность':<12} {'Вход':<12} {'Стоп':<12} {'Цель 1':<12} {'R/R':<8}")
        print("-"*80)
        
        for i, signal in enumerate(strong_signals, 1):
            levels = signal['levels']
            print(f"{i:<4} {signal['symbol']:<15} {signal['confidence']:<12} "
                  f"{levels['entry_price']:<12.6f} {levels['stop_loss']:<12.6f} "
                  f"{levels['take_profit_1']:<12.6f} {levels['rr_ratio_1']:<8.2f}")
        
        # Экспорт в Excel
        print(f"\n{'='*80}")
        print("📊 ЭКСПОРТ РЕЗУЛЬТАТОВ В EXCEL")
        print(f"{'='*80}")
        finder.export_to_excel(strong_signals)
        
    else:
        print("\n❌ Сильных сигналов для шорта не найдено")
        print("   💡 Рекомендации:")
        print("      • Попробуйте позже, когда рынок будет в перекупленности")
        print("      • Проверьте другие таймфреймы")
        print("      • Снизьте требования к минимальной уверенности (не рекомендуется)")
    
    print(f"\n⏰ Конец анализа: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*80}\n")
    
    return strong_signals


def test_strong_short_finder():
    """Функция для тестирования"""
    try:
        results = main_strong_short_search()
        success = len(results) > 0
        print(f"\n{'✅' if success else '⚠️'} Поиск завершен! Найдено {len(results)} сильных сигналов")
        return success
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    test_strong_short_finder()
