import time
import os
import numpy as np
import pandas as pd
import requests
import talib
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl.styles import PatternFill, Font, Alignment


def get_high_volume_symbols(min_volume=10_000_000):
    """Получает список монет с высоким объемом (MEXC futures)."""
    url = "https://contract.mexc.com/api/v1/contract/ticker"
    try:
        response = requests.get(url, timeout=10)
        data = response.json()
        symbols = []
        if "data" in data:
            for item in data["data"]:
                if item.get("amount24", 0) > min_volume:
                    symbols.append({
                        "symbol": item["symbol"],
                        "volume_24h": item["amount24"],
                        "last_price": float(item["lastPrice"])
                    })
        return sorted(symbols, key=lambda x: x["volume_24h"], reverse=True)
    except Exception as e:
        print(f"Ошибка получения списка монет: {e}")
        return []


class VSAHourlyShortFinder:
    """
    Поиск SHORT/LONG сигналов по Volume Price Analysis (VPA / VSA)
    на таймфрейме 30 минут (MEXC). Оптимизировано под ожидание разворота,
    без жесткого ограничения времени сделки.
    """

    def __init__(self):
        self.min_score = 70
        self.min_rr_ratio = 2.0
        self.min_volume_24h = 10_000_000
        self.debug = True

    def get_candles(self, symbol, interval="Min30", limit=200):
        url = f"https://contract.mexc.com/api/v1/contract/kline/{symbol}"
        params = {"interval": interval, "limit": limit}
        try:
            response = requests.get(url, params=params, timeout=15)
            if response.status_code == 200:
                data = response.json()
                if self.debug and (not data or not data.get("success")):
                    print(f"   ⚠️ Пустой ответ API для {symbol}: {data}")
                return data
        except Exception as e:
            print(f"   ⚠️ Ошибка получения данных для {symbol}: {e}")
        return None

    def create_dataframe(self, data):
        if not data or not data.get("success") or not data.get("data"):
            return None

        raw = data["data"]
        try:
            if isinstance(raw, dict):
                required = ["time", "open", "close", "high", "low", "vol"]
                if not all(k in raw for k in required):
                    return None
                # Нормализуем длины массивов
                length = min(
                    len(raw["time"]),
                    len(raw["open"]),
                    len(raw["high"]),
                    len(raw["low"]),
                    len(raw["close"]),
                    len(raw["vol"]),
                )
                if length == 0:
                    return None
                df = pd.DataFrame({
                    "timestamp": raw["time"][:length],
                    "open": raw["open"][:length],
                    "high": raw["high"][:length],
                    "low": raw["low"][:length],
                    "close": raw["close"][:length],
                    "volume": raw["vol"][:length],
                })
            elif isinstance(raw, list):
                if len(raw) == 0:
                    return None
                # Поддержка списка списков с разной длиной строк
                cleaned = []
                for row in raw:
                    if isinstance(row, (list, tuple)) and len(row) >= 6:
                        cleaned.append(row[:6])
                if not cleaned:
                    return None
                df = pd.DataFrame(cleaned, columns=["timestamp", "open", "high", "low", "close", "volume"])
            else:
                return None

            for col in ["open", "high", "low", "close", "volume"]:
                df[col] = pd.to_numeric(df[col], errors="coerce")
            df = df.dropna()
            if len(df) < 80:
                return None

            # Корректная конвертация timestamp (сек/мс)
            ts = df["timestamp"].iloc[0]
            if ts > 1e12:
                df["datetime"] = pd.to_datetime(df["timestamp"], unit="ms")
            else:
                df["datetime"] = pd.to_datetime(df["timestamp"], unit="s")
            df = df.sort_values("datetime").reset_index(drop=True)
            if self.debug:
                first_ts = df["datetime"].iloc[0]
                last_ts = df["datetime"].iloc[-1]
                print(f"   📅 Свечи: {len(df)} | {first_ts} → {last_ts}")
            return df
        except Exception as e:
            print(f"   ⚠️ Ошибка создания DataFrame: {e}")
            return None

    def add_indicators(self, df):
        close = df["close"].values
        high = df["high"].values
        low = df["low"].values
        volume = df["volume"].values

        df["sma_50"] = talib.SMA(close, timeperiod=50)
        df["sma_200"] = talib.SMA(close, timeperiod=200)
        df["atr"] = talib.ATR(high, low, close, timeperiod=14)
        df["volume_sma"] = talib.SMA(volume, timeperiod=20)
        df["spread"] = df["high"] - df["low"]
        df["avg_spread"] = df["spread"].rolling(20).mean()
        df["close_pos"] = (df["close"] - df["low"]) / (df["spread"].replace(0, np.nan))
        df["is_up"] = df["close"] > df["open"]
        df["is_down"] = df["close"] < df["open"]
        df["resistance"] = df["high"].rolling(20).max()
        df["support"] = df["low"].rolling(20).min()
        return df

    def detect_vsa_signals(self, df):
        """Возвращает сигнал и баллы по VSA (SHORT/LONG)."""
        if len(df) < 80:
            return None

        # Берем последнюю ЗАКРЫТУЮ свечу как сигнальную
        signal_candle = df.iloc[-2]
        confirm_candle = df.iloc[-1]

        signals = []
        score = 0
        direction = None

        vol_ratio = signal_candle["volume"] / signal_candle["volume_sma"] if signal_candle["volume_sma"] else 0
        spread_ratio = signal_candle["spread"] / signal_candle["avg_spread"] if signal_candle["avg_spread"] else 0
        close_pos = signal_candle["close_pos"]
        close = signal_candle["close"]
        sma_50 = df["sma_50"].iloc[-2]
        sma_200 = df["sma_200"].iloc[-2] if not pd.isna(df["sma_200"].iloc[-2]) else sma_50

        # 1) Upthrust (ложный пробой вверх на высоком объеме)
        if (
            signal_candle["is_up"]
            and vol_ratio > 1.5
            and spread_ratio > 1.5
            and close_pos < 0.35
        ):
            signals.append("Upthrust: широкий спред вверх, закрытие у низа, высокий объем")
            score += 30
            direction = "SHORT"

        # 2) No Demand (слабый рост на низком объеме)
        if (
            signal_candle["is_up"]
            and vol_ratio < 0.7
            and spread_ratio < 0.8
            and close_pos > 0.6
        ):
            signals.append("No Demand: рост на низком объеме и узком спреде")
            score += 20
            direction = "SHORT"

        # 3) Effort vs Result (много объема, мало результата)
        if vol_ratio > 1.8 and spread_ratio < 0.9:
            signals.append("Effort vs Result: высокий объем при малом спреде")
            score += 20
            direction = direction or "SHORT"

        # 4) Climactic Volume near top (климактический объем у вершины)
        if vol_ratio > 2.2 and close_pos < 0.4 and signal_candle["close"] > df["sma_50"].iloc[-2]:
            signals.append("Climactic Volume: всплеск объема у вершины")
            score += 25
            direction = "SHORT"

        # LONG сигналы по VSA
        # 5) Spring (ложный пробой вниз на высоком объеме)
        if (
            signal_candle["is_down"]
            and vol_ratio > 1.5
            and spread_ratio > 1.5
            and close_pos > 0.65
        ):
            signals.append("Spring: широкий спред вниз, закрытие у верха, высокий объем")
            score += 30
            direction = "LONG"

        # 6) No Supply (слабое снижение на низком объеме)
        if (
            signal_candle["is_down"]
            and vol_ratio < 0.7
            and spread_ratio < 0.8
            and close_pos < 0.4
        ):
            signals.append("No Supply: снижение на низком объеме и узком спреде")
            score += 20
            direction = "LONG"

        # 7) Climactic Volume near bottom (климактический объем у дна)
        if vol_ratio > 2.2 and close_pos > 0.6 and signal_candle["close"] < df["sma_50"].iloc[-2]:
            signals.append("Climactic Volume: всплеск объема у дна")
            score += 25
            direction = "LONG"

        # Подтверждение следующей свечой (должна закрыться ниже)
        if direction == "SHORT" and confirm_candle["close"] < signal_candle["close"]:
            signals.append("Подтверждение: следующая свеча закрылась ниже")
            score += 10
        elif direction == "LONG" and confirm_candle["close"] > signal_candle["close"]:
            signals.append("Подтверждение: следующая свеча закрылась выше")
            score += 10
        else:
            score -= 10

        # Фильтр: сильный объем
        if vol_ratio < 0.8:
            score -= 10

        # Штраф за слишком узкий спред (сигнал слабее)
        if spread_ratio < 0.7:
            signals.append("⚠️ Штраф: слишком узкий спред относительно среднего")
            score -= 15

        # Локальная зона (поддержка/сопротивление)
        resistance = df["resistance"].iloc[-2]
        support = df["support"].iloc[-2]
        if direction == "SHORT":
            dist_to_res = abs(resistance - close) / close if close else 0
            if dist_to_res <= 0.01:
                signals.append("Локальная зона: цена близко к сопротивлению (≤1%)")
                score += 10
        elif direction == "LONG":
            dist_to_sup = abs(close - support) / close if close else 0
            if dist_to_sup <= 0.01:
                signals.append("Локальная зона: цена близко к поддержке (≤1%)")
                score += 10

        # Трендовый фильтр: торгуем по направлению старшего тренда
        if direction == "SHORT" and close < sma_50 and sma_50 < sma_200:
            signals.append("Тренд подтвержден: цена ниже SMA50 и SMA200")
            score += 10
        elif direction == "LONG" and close > sma_50 and sma_50 > sma_200:
            signals.append("Тренд подтвержден: цена выше SMA50 и SMA200")
            score += 10
        else:
            score -= 5

        if score <= 0:
            return None

        return {
            "signals": signals,
            "score": min(score, 100),
            "direction": direction or "SHORT",
            "signal_time": signal_candle["datetime"],
            "vol_ratio": round(vol_ratio, 2),
            "spread_ratio": round(spread_ratio, 2),
            "close_pos": round(close_pos, 2),
        }

    def detect_vsa_candidates(self, df):
        """Ослабленные VSA-кандидаты для watchlist."""
        if len(df) < 80:
            return None

        signal_candle = df.iloc[-2]
        confirm_candle = df.iloc[-1]

        signals = []
        score = 0
        direction = None

        vol_ratio = signal_candle["volume"] / signal_candle["volume_sma"] if signal_candle["volume_sma"] else 0
        spread_ratio = signal_candle["spread"] / signal_candle["avg_spread"] if signal_candle["avg_spread"] else 0
        close_pos = signal_candle["close_pos"]

        # Ослабленные условия SHORT
        if signal_candle["is_up"] and vol_ratio > 1.2 and spread_ratio > 1.2 and close_pos < 0.45:
            signals.append("Кандидат Upthrust (ослабленные условия)")
            score += 18
            direction = "SHORT"

        if signal_candle["is_up"] and vol_ratio < 0.9 and spread_ratio < 1.0 and close_pos > 0.55:
            signals.append("Кандидат No Demand (ослабленные условия)")
            score += 12
            direction = "SHORT"

        # Ослабленные условия LONG
        if signal_candle["is_down"] and vol_ratio > 1.2 and spread_ratio > 1.2 and close_pos > 0.55:
            signals.append("Кандидат Spring (ослабленные условия)")
            score += 18
            direction = "LONG"

        if signal_candle["is_down"] and vol_ratio < 0.9 and spread_ratio < 1.0 and close_pos < 0.45:
            signals.append("Кандидат No Supply (ослабленные условия)")
            score += 12
            direction = "LONG"

        # Усиливающие факторы
        if vol_ratio > 1.4 and spread_ratio < 1.0:
            signals.append("Кандидат Effort vs Result (ослабленные условия)")
            score += 10
            direction = direction or "SHORT"

        if confirm_candle["close"] < signal_candle["close"] and direction == "SHORT":
            signals.append("Подтверждение вниз (слабое)")
            score += 6
        elif confirm_candle["close"] > signal_candle["close"] and direction == "LONG":
            signals.append("Подтверждение вверх (слабое)")
            score += 6

        if score <= 0:
            return None

        return {
            "signals": signals,
            "score": min(score, 60),
            "direction": direction or "SHORT",
            "signal_time": signal_candle["datetime"],
            "vol_ratio": round(vol_ratio, 2),
            "spread_ratio": round(spread_ratio, 2),
            "close_pos": round(close_pos, 2),
        }

    def calculate_trade_levels(self, df, direction):
        """Уровни для сделки по развороту (без жесткого лимита времени)."""
        signal_candle = df.iloc[-2]
        current_price = df.iloc[-1]["close"]
        atr = df["atr"].iloc[-2]
        if pd.isna(atr) or atr == 0:
            atr = current_price * 0.01

        if direction == "SHORT":
            stop_loss = max(signal_candle["high"] * 1.002, current_price + atr * 0.8)
            take_profit_1 = current_price - atr * 1.2
            take_profit_2 = current_price - atr * 1.6
            take_profit_3 = current_price - atr * 2.0
            risk = stop_loss - current_price
            reward_1 = current_price - take_profit_1
        else:
            stop_loss = min(signal_candle["low"] * 0.998, current_price - atr * 0.8)
            take_profit_1 = current_price + atr * 1.2
            take_profit_2 = current_price + atr * 1.6
            take_profit_3 = current_price + atr * 2.0
            risk = current_price - stop_loss
            reward_1 = take_profit_1 - current_price

        rr_ratio_1 = reward_1 / risk if risk > 0 else 0

        # Оценка длительности сделки (ориентировочно)
        avg_spread = df["avg_spread"].iloc[-2]
        if pd.isna(avg_spread) or avg_spread == 0:
            avg_spread = atr
        hours_to_target = max(1.0, min(12.0, abs(take_profit_1 - current_price) / avg_spread))
        max_hold_hours = max(4.0, min(24.0, hours_to_target * 2.0))

        return {
            "entry_price": round(current_price, 6),
            "stop_loss": round(stop_loss, 6),
            "take_profit_1": round(take_profit_1, 6),
            "take_profit_2": round(take_profit_2, 6),
            "take_profit_3": round(take_profit_3, 6),
            "rr_ratio_1": round(rr_ratio_1, 2),
            "atr": round(atr, 6),
            "expected_hold_hours": round(hours_to_target, 1),
            "max_hold_hours": round(max_hold_hours, 1),
        }

    def analyze_symbol(self, symbol):
        print(f"\n🔍 Анализируем {symbol} (VSA, 30m) ...")
        data = self.get_candles(symbol, "Min30", 200)
        if not data:
            print(f"   ❌ Нет данных для {symbol}")
            return None, None

        df = self.create_dataframe(data)
        if df is None:
            print(f"   ❌ Не удалось создать данные для {symbol}")
            return None, None

        df = self.add_indicators(df)
        vsa = self.detect_vsa_signals(df)
        if not vsa:
            print(f"   ⚠️ Нет четких VSA сигналов для {symbol}")
            # Пробуем ослабленные кандидаты для watchlist
            candidate = self.detect_vsa_candidates(df)
            if not candidate:
                return None, None
            vsa = candidate

        levels = self.calculate_trade_levels(df, vsa["direction"])
        chart_path = ""
        scenario = self.build_scenario_text(vsa["direction"], levels)

        result = {
            "symbol": symbol,
            "score": vsa["score"],
            "direction": vsa["direction"],
            "signal_time": vsa["signal_time"].strftime("%Y-%m-%d %H:%M:%S"),
            "signals": vsa["signals"],
            "levels": levels,
            "vol_ratio": vsa["vol_ratio"],
            "spread_ratio": vsa["spread_ratio"],
            "close_pos": vsa["close_pos"],
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "chart_path": chart_path,
            "scenario": scenario,
        }

        # Основной сигнал
        if vsa["score"] >= self.min_score and levels["rr_ratio_1"] >= self.min_rr_ratio:
            result["chart_path"] = self.save_chart(df, symbol, vsa["direction"], levels)
            return result, None

        # Watchlist (почти сигнал)
        watch_reason = []
        if vsa["score"] < self.min_score:
            watch_reason.append(f"Низкий балл {vsa['score']}/100")
        if levels["rr_ratio_1"] < self.min_rr_ratio:
            watch_reason.append(f"Слабое R/R {levels['rr_ratio_1']}:1")

        if vsa["score"] >= 25:
            watch_hint = []
            if vsa["score"] < self.min_score:
                watch_hint.append("Нужна более сильная комбинация VSA (объем/спред/подтверждение)")
            if levels["rr_ratio_1"] < self.min_rr_ratio:
                watch_hint.append("Нужен больший потенциал хода или более короткий стоп")

            if vsa["direction"] == "SHORT":
                watch_hint.append("Развитие вниз вероятно при пробое минимума сигнальной свечи")
                watch_hint.append("Отмена сценария — закрепление выше входа")
            else:
                watch_hint.append("Развитие вверх вероятно при пробое максимума сигнальной свечи")
                watch_hint.append("Отмена сценария — закрепление ниже входа")

            result["watch_reason"] = " | ".join(watch_reason) if watch_reason else "Нужна доп.подтверждение"
            result["watch_hint"] = " | ".join(watch_hint)
            return None, result

        return None, None

    def build_scenario_text(self, direction, levels):
        """Текст сценариев развития события."""
        if direction == "SHORT":
            return (
                f"Базовый: пробой вниз и движение к цели 1 ({levels['take_profit_1']}). "
                f"Альтернатива: возврат выше входа и закрепление → стоп ({levels['stop_loss']})."
            )
        return (
            f"Базовый: рост к цели 1 ({levels['take_profit_1']}). "
            f"Альтернатива: возврат ниже входа и закрепление → стоп ({levels['stop_loss']})."
        )

    def save_chart(self, df, symbol, direction, levels, out_dir="charts_vsa"):
        """Сохраняет график с точкой входа и целями."""
        try:
            os.makedirs(out_dir, exist_ok=True)
            tail = df.tail(120).copy()
            x = range(len(tail))
            prices = tail["close"].values

            plt.figure(figsize=(12, 6))
            plt.plot(x, prices, label="Цена (close)", linewidth=1.5)
            plt.axhline(levels["entry_price"], color="blue", linestyle="--", label="Вход")
            plt.axhline(levels["stop_loss"], color="red", linestyle="--", label="Стоп")
            plt.axhline(levels["take_profit_1"], color="green", linestyle="--", label="Цель 1")
            plt.axhline(levels["take_profit_2"], color="orange", linestyle="--", label="Цель 2")
            plt.axhline(levels["take_profit_3"], color="purple", linestyle="--", label="Цель 3")

            title = f"{symbol} | {direction} | 30m VSA"
            plt.title(title)
            plt.legend()
            plt.tight_layout()

            filename = os.path.join(out_dir, f"{symbol}_{direction}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png")
            plt.savefig(filename, dpi=150)
            plt.close()
            return os.path.abspath(filename)
        except Exception:
            return ""
    def export_to_excel(self, results, watchlist, filename="vsa_signals.xlsx"):
        # Формируем пустые таблицы с заголовками, чтобы файл создавался всегда
        signal_columns = [
            "№", "Символ", "Сила сигнала", "Направление", "Время сигнала", "Время анализа",
            "Сигналы VSA", "Объем/средний", "Спред/средний", "Позиция закрытия",
            "Вход", "Стоп", "Цель 1", "Цель 2", "Цель 3", "R/R", "ATR",
            "Ожид. часы до цели", "Макс. удержание (часы)", "Сценарий", "График"
        ]
        watch_columns = [
            "№", "Символ", "Сила сигнала", "Направление", "Время сигнала", "Время анализа",
            "Причина в watchlist", "Комментарий по развитию", "Сигналы VSA", "Объем/средний",
            "Спред/средний", "Позиция закрытия", "Вход", "Стоп", "Цель 1", "Цель 2",
            "Цель 3", "R/R", "ATR", "Ожид. часы до цели", "Макс. удержание (часы)"
        ]

        rows = []
        for i, r in enumerate(results, 1):
            levels = r["levels"]
            rows.append({
                "№": i,
                "Символ": r["symbol"],
                "Сила сигнала": r["score"],
                "Направление": r["direction"],
                "Время сигнала": r["signal_time"],
                "Время анализа": r["timestamp"],
                "Сигналы VSA": " | ".join(r["signals"]),
                "Объем/средний": r["vol_ratio"],
                "Спред/средний": r["spread_ratio"],
                "Позиция закрытия": r["close_pos"],
                "Вход": levels["entry_price"],
                "Стоп": levels["stop_loss"],
                "Цель 1": levels["take_profit_1"],
                "Цель 2": levels["take_profit_2"],
                "Цель 3": levels["take_profit_3"],
                "R/R": levels["rr_ratio_1"],
                "ATR": levels["atr"],
                "Ожид. часы до цели": levels["expected_hold_hours"],
                "Макс. удержание (часы)": levels["max_hold_hours"],
                "Сценарий": r["scenario"],
                "График": r["chart_path"],
            })

        df = pd.DataFrame(rows, columns=signal_columns)

        watch_rows = []
        for i, r in enumerate(watchlist, 1):
            levels = r["levels"]
            watch_rows.append({
                "№": i,
                "Символ": r["symbol"],
                "Сила сигнала": r["score"],
                "Направление": r["direction"],
                "Время сигнала": r["signal_time"],
                "Время анализа": r["timestamp"],
                "Причина в watchlist": r.get("watch_reason", ""),
                "Комментарий по развитию": r.get("watch_hint", ""),
                "Сигналы VSA": " | ".join(r["signals"]),
                "Объем/средний": r["vol_ratio"],
                "Спред/средний": r["spread_ratio"],
                "Позиция закрытия": r["close_pos"],
                "Вход": levels["entry_price"],
                "Стоп": levels["stop_loss"],
                "Цель 1": levels["take_profit_1"],
                "Цель 2": levels["take_profit_2"],
                "Цель 3": levels["take_profit_3"],
                "R/R": levels["rr_ratio_1"],
                "ATR": levels["atr"],
                "Ожид. часы до цели": levels["expected_hold_hours"],
                "Макс. удержание (часы)": levels["max_hold_hours"],
            })

        try:
            with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
                if not df.empty:
                    df.to_excel(writer, sheet_name="Сигналы", index=False)
                if watch_rows:
                    df_watch = pd.DataFrame(watch_rows, columns=watch_columns)
                    df_watch.to_excel(writer, sheet_name="Watchlist", index=False)
                elif df.empty:
                    # Если нет данных вообще, создаем пустой Watchlist для наглядности
                    df_watch = pd.DataFrame([], columns=watch_columns)
                    df_watch.to_excel(writer, sheet_name="Watchlist", index=False)

                # Авто-ширина и оформление
                for sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    for column in ws.columns:
                        max_length = 0
                        col_letter = column[0].column_letter
                        for cell in column:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        ws.column_dimensions[col_letter].width = min(max_length + 2, 100)

                    # Выделяем Цель 2 красным
                    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    bold_font = Font(bold=True, color="CC0000")
                    center_alignment = Alignment(horizontal="center", vertical="center")
                    header_fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")

                    for col_idx, cell in enumerate(ws[1], start=1):
                        if cell.value and "Цель 2" in str(cell.value):
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = center_alignment
                            for row_idx in range(2, ws.max_row + 1):
                                c = ws.cell(row=row_idx, column=col_idx)
                                c.fill = red_fill
                                c.font = bold_font
                                c.alignment = center_alignment

                    # Выделение watchlist другим цветом
                    if sheet_name == "Watchlist":
                        watch_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        for row_idx in range(2, ws.max_row + 1):
                            for col_idx in range(1, ws.max_column + 1):
                                ws.cell(row=row_idx, column=col_idx).fill = watch_fill

            print(f"\n✅ Excel сохранен: {filename}")
            print(f"📁 Путь: {os.path.abspath(filename)}")
            return filename
        except Exception as e:
            print(f"\n❌ Ошибка экспорта в Excel: {e}")
            return None


def main():
    print("🚀 VSA СИГНАЛЫ (SHORT/LONG) НА ЧАСОВОМ ТАЙМФРЕЙМЕ")
    print("================================================================================")
    print("Основано на принципах Anna Coulling (VPA/VSA): объем + спред + позиция закрытия")
    print("Таймфрейм: 30 минут | Фокус: разворот и удержание в тренде")
    print("================================================================================")

    finder = VSAHourlyShortFinder()
    symbols_data = get_high_volume_symbols(min_volume=finder.min_volume_24h)
    if not symbols_data:
        print("❌ Не удалось получить список монет")
        return []

    symbols = [s["symbol"] for s in symbols_data]
    print(f"\n📊 Анализируем {len(symbols)} монет (объем > 10 млн)...")

    results = []
    watchlist = []
    for symbol in symbols:
        res, watch = finder.analyze_symbol(symbol)
        if res:
            results.append(res)
        if watch:
            watchlist.append(watch)
        time.sleep(0.4)

    results.sort(key=lambda x: x["score"], reverse=True)
    watchlist.sort(key=lambda x: x["score"], reverse=True)
    if results:
        print(f"\n✅ Найдено сигналов: {len(results)}")
    if watchlist:
        print(f"📌 Watchlist: {len(watchlist)} кандидатов")
    if not results and not watchlist:
        print("\n❌ Сильных VSA сигналов не найдено")
        print("📝 Файл будет создан пустым, чтобы не искать его вручную.")

    finder.export_to_excel(results, watchlist)

    return results


if __name__ == "__main__":
    main()
